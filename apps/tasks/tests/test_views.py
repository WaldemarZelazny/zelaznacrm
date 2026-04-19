"""Testy widokow CRUD i kalendarza aplikacji tasks."""

from __future__ import annotations

import datetime
import json

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from apps.accounts.models import UserProfile
from apps.companies.models import Company
from apps.deals.models import Deal
from apps.leads.models import Lead, WorkflowStage
from apps.tasks.models import Task

# ---------------------------------------------------------------------------
# Pomocnicze funkcje
# ---------------------------------------------------------------------------


def _make_user(username: str, role: str = UserProfile.Role.HANDLOWIEC) -> User:
    """Tworzy uzytkownika z profilem o podanej roli."""
    user = User.objects.create_user(username=username, password="testpass")
    user.profile.role = role
    user.profile.save()
    return user


def _make_company(name: str, owner: User | None = None) -> Company:
    """Tworzy firme testowa."""
    return Company.objects.create(name=name, owner=owner)


def _make_task(
    title: str,
    assigned_to: User | None = None,
    company: Company | None = None,
    status: str = Task.Status.DO_ZROBIENIA,
    priority: str = Task.Priority.SREDNI,
    due_offset_hours: int = 24,
) -> Task:
    """Tworzy zadanie testowe."""
    return Task.objects.create(
        title=title,
        assigned_to=assigned_to,
        company=company,
        status=status,
        priority=priority,
        due_date=timezone.now() + datetime.timedelta(hours=due_offset_hours),
    )


def _make_lead(title: str, company: Company, owner: User | None = None) -> Lead:
    """Tworzy lead testowy."""
    stage, _ = WorkflowStage.objects.get_or_create(
        order=1, defaults={"name": "Nowy", "color": "#6c757d"}
    )
    return Lead.objects.create(title=title, company=company, owner=owner, stage=stage)


def _make_deal(title: str, company: Company, owner: User | None = None) -> Deal:
    """Tworzy umowe testowa."""
    return Deal.objects.create(
        title=title,
        company=company,
        owner=owner,
        value="5000.00",
        close_date=datetime.date.today() + datetime.timedelta(days=30),
    )


# ---------------------------------------------------------------------------
# Testy: przekierowanie niezalogowanych
# ---------------------------------------------------------------------------


class TaskViewsAuthRedirectTest(TestCase):
    """Niezalogowani uzytkowniczy musza byc przekierowani na login."""

    def setUp(self) -> None:
        self.user = _make_user("owner_user")
        self.company = _make_company("Acme SA", owner=self.user)
        self.task = _make_task("Test task", assigned_to=self.user, company=self.company)

    def _assert_redirects_to_login(self, url: str, method: str = "get") -> None:
        response = getattr(self.client, method)(url)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_list_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("tasks:list"))

    def test_calendar_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("tasks:calendar"))

    def test_detail_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("tasks:detail", kwargs={"pk": self.task.pk})
        )

    def test_create_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("tasks:create"))

    def test_update_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("tasks:update", kwargs={"pk": self.task.pk})
        )

    def test_delete_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("tasks:delete", kwargs={"pk": self.task.pk})
        )

    def test_complete_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("tasks:complete", kwargs={"pk": self.task.pk}), method="post"
        )

    def test_cancel_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("tasks:cancel", kwargs={"pk": self.task.pk}), method="post"
        )


# ---------------------------------------------------------------------------
# Testy: TaskListView
# ---------------------------------------------------------------------------


class TaskListViewTest(TestCase):
    """Testy widoku listy zadan."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", role=UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec_user")
        self.company = _make_company("Acme SA", owner=self.handlowiec)
        self.own_task = _make_task(
            "Moje zadanie", assigned_to=self.handlowiec, company=self.company
        )
        self.other_task = _make_task(
            "Cudze zadanie", assigned_to=self.admin, company=self.company
        )

    def test_list_returns_200(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:list"))
        self.assertEqual(response.status_code, 200)

    def test_handlowiec_sees_only_own_tasks(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:list"))
        tasks = list(response.context["tasks"])
        self.assertIn(self.own_task, tasks)
        self.assertNotIn(self.other_task, tasks)

    def test_admin_sees_all_tasks(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:list"))
        tasks = list(response.context["tasks"])
        self.assertIn(self.own_task, tasks)
        self.assertIn(self.other_task, tasks)

    def test_filter_by_status(self) -> None:
        done_task = _make_task(
            "Wykonane",
            assigned_to=self.handlowiec,
            company=self.company,
            status=Task.Status.WYKONANE,
        )
        self.client.force_login(self.handlowiec)
        response = self.client.get(
            reverse("tasks:list"), {"status": Task.Status.WYKONANE}
        )
        tasks = list(response.context["tasks"])
        self.assertIn(done_task, tasks)
        self.assertNotIn(self.own_task, tasks)

    def test_filter_by_priority(self) -> None:
        urgent = _make_task(
            "Pilne",
            assigned_to=self.handlowiec,
            company=self.company,
            priority=Task.Priority.PILNY,
        )
        self.client.force_login(self.handlowiec)
        response = self.client.get(
            reverse("tasks:list"), {"priority": Task.Priority.PILNY}
        )
        tasks = list(response.context["tasks"])
        self.assertIn(urgent, tasks)
        self.assertNotIn(self.own_task, tasks)

    def test_filter_by_task_type(self) -> None:
        call_task = _make_task(
            "Telefon",
            assigned_to=self.handlowiec,
            company=self.company,
        )
        call_task.task_type = Task.TaskType.TELEFON
        call_task.save()
        self.client.force_login(self.handlowiec)
        response = self.client.get(
            reverse("tasks:list"), {"task_type": Task.TaskType.TELEFON}
        )
        tasks = list(response.context["tasks"])
        self.assertIn(call_task, tasks)
        self.assertNotIn(self.own_task, tasks)

    def test_context_contains_filter_choices(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:list"))
        self.assertIn("status_choices", response.context)
        self.assertIn("priority_choices", response.context)
        self.assertIn("task_type_choices", response.context)


# ---------------------------------------------------------------------------
# Testy: TaskCalendarView — HTML i JSON
# ---------------------------------------------------------------------------


class TaskCalendarViewTest(TestCase):
    """Testy widoku kalendarza — strona HTML oraz JSON endpoint."""

    def setUp(self) -> None:
        self.user = _make_user("cal_user")
        self.company = _make_company("Test Co", owner=self.user)
        self.task = _make_task("Spotkanie", assigned_to=self.user, company=self.company)

    def test_calendar_html_returns_200(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:calendar"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "tasks/task_calendar.html")

    def test_calendar_json_returns_200(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:calendar"), {"format": "json"})
        self.assertEqual(response.status_code, 200)

    def test_calendar_json_content_type(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:calendar"), {"format": "json"})
        self.assertEqual(response["Content-Type"], "application/json")

    def test_calendar_json_structure(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:calendar"), {"format": "json"})
        data = json.loads(response.content)
        self.assertIsInstance(data, list)
        self.assertEqual(len(data), 1)
        event = data[0]
        self.assertIn("id", event)
        self.assertIn("title", event)
        self.assertIn("start", event)
        self.assertIn("color", event)
        self.assertIn("url", event)
        self.assertIn("extendedProps", event)

    def test_calendar_json_extended_props(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:calendar"), {"format": "json"})
        data = json.loads(response.content)
        props = data[0]["extendedProps"]
        self.assertIn("status", props)
        self.assertIn("priority", props)

    def test_handlowiec_json_sees_only_own_tasks(self) -> None:
        other = _make_user("other_cal")
        _make_task("Cudze", assigned_to=other, company=self.company)
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:calendar"), {"format": "json"})
        data = json.loads(response.content)
        ids = [event["id"] for event in data]
        self.assertIn(self.task.pk, ids)
        self.assertEqual(len(ids), 1)

    def test_admin_json_sees_all_tasks(self) -> None:
        admin = _make_user("admin_cal", role=UserProfile.Role.ADMIN)
        other = _make_user("other_for_admin")
        _make_task("Cudze dla admina", assigned_to=other, company=self.company)
        self.client.force_login(admin)
        response = self.client.get(reverse("tasks:calendar"), {"format": "json"})
        data = json.loads(response.content)
        self.assertGreaterEqual(len(data), 2)


# ---------------------------------------------------------------------------
# Testy: TaskDetailView
# ---------------------------------------------------------------------------


class TaskDetailViewTest(TestCase):
    """Testy widoku szczegolowego zadania."""

    def setUp(self) -> None:
        self.owner = _make_user("detail_owner")
        self.other = _make_user("detail_other")
        self.admin = _make_user("detail_admin", role=UserProfile.Role.ADMIN)
        self.company = _make_company("Detail Co", owner=self.owner)
        self.task = _make_task(
            "Detail task", assigned_to=self.owner, company=self.company
        )

    def test_owner_can_view_task(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("tasks:detail", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 200)

    def test_admin_can_view_any_task(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:detail", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 200)

    def test_other_handlowiec_gets_404(self) -> None:
        self.client.force_login(self.other)
        response = self.client.get(reverse("tasks:detail", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 404)

    def test_context_has_can_edit_true_for_owner(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("tasks:detail", kwargs={"pk": self.task.pk}))
        self.assertTrue(response.context["can_edit"])

    def test_context_has_is_admin_for_admin(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:detail", kwargs={"pk": self.task.pk}))
        self.assertTrue(response.context["is_admin"])


# ---------------------------------------------------------------------------
# Testy: TaskCreateView
# ---------------------------------------------------------------------------


class TaskCreateViewTest(TestCase):
    """Testy widoku tworzenia zadania."""

    def setUp(self) -> None:
        self.user = _make_user("create_user")
        self.company = _make_company("Create Co", owner=self.user)

    def _due_date_str(self, offset_days: int = 1) -> str:
        dt = timezone.now() + datetime.timedelta(days=offset_days)
        return dt.strftime("%Y-%m-%dT%H:%M")

    def test_create_get_returns_200(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:create"))
        self.assertEqual(response.status_code, 200)

    def test_create_task_success(self) -> None:
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("tasks:create"),
            {
                "title": "Nowe zadanie",
                "task_type": Task.TaskType.ZADANIE,
                "priority": Task.Priority.SREDNI,
                "status": Task.Status.DO_ZROBIENIA,
                "due_date": self._due_date_str(),
            },
        )
        self.assertEqual(Task.objects.filter(title="Nowe zadanie").count(), 1)
        task = Task.objects.get(title="Nowe zadanie")
        self.assertRedirects(response, reverse("tasks:detail", kwargs={"pk": task.pk}))

    def test_create_sets_created_by(self) -> None:
        self.client.force_login(self.user)
        self.client.post(
            reverse("tasks:create"),
            {
                "title": "Zadanie z created_by",
                "task_type": Task.TaskType.EMAIL,
                "priority": Task.Priority.NISKI,
                "status": Task.Status.DO_ZROBIENIA,
                "due_date": self._due_date_str(),
            },
        )
        task = Task.objects.get(title="Zadanie z created_by")
        self.assertEqual(task.created_by, self.user)

    def test_create_prefill_company_id(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(
            reverse("tasks:create"), {"company_id": self.company.pk}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["form"].initial.get("company"), self.company)

    def test_create_prefill_lead_id(self) -> None:
        lead = _make_lead("Lead prefill", self.company, owner=self.user)
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:create"), {"lead_id": lead.pk})
        self.assertEqual(response.status_code, 200)
        initial = response.context["form"].initial
        self.assertEqual(initial.get("lead"), lead)
        self.assertEqual(initial.get("company"), self.company)

    def test_create_prefill_deal_id(self) -> None:
        deal = _make_deal("Deal prefill", self.company, owner=self.user)
        self.client.force_login(self.user)
        response = self.client.get(reverse("tasks:create"), {"deal_id": deal.pk})
        self.assertEqual(response.status_code, 200)
        initial = response.context["form"].initial
        self.assertEqual(initial.get("deal"), deal)
        self.assertEqual(initial.get("company"), self.company)


# ---------------------------------------------------------------------------
# Testy: TaskUpdateView
# ---------------------------------------------------------------------------


class TaskUpdateViewTest(TestCase):
    """Testy widoku edycji zadania."""

    def setUp(self) -> None:
        self.owner = _make_user("update_owner")
        self.other = _make_user("update_other")
        self.admin = _make_user("update_admin", role=UserProfile.Role.ADMIN)
        self.company = _make_company("Update Co", owner=self.owner)
        self.task = _make_task(
            "Update task", assigned_to=self.owner, company=self.company
        )

    def _due_date_str(self, offset_days: int = 1) -> str:
        dt = timezone.now() + datetime.timedelta(days=offset_days)
        return dt.strftime("%Y-%m-%dT%H:%M")

    def test_owner_can_edit(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("tasks:update", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 200)

    def test_other_handlowiec_gets_403(self) -> None:
        self.client.force_login(self.other)
        response = self.client.get(reverse("tasks:update", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 403)

    def test_admin_can_edit_any_task(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:update", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 200)

    def test_update_task_success(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse("tasks:update", kwargs={"pk": self.task.pk}),
            {
                "title": "Zaktualizowane zadanie",
                "task_type": Task.TaskType.EMAIL,
                "priority": Task.Priority.WYSOKI,
                "status": Task.Status.W_TOKU,
                "due_date": self._due_date_str(),
            },
        )
        self.task.refresh_from_db()
        self.assertEqual(self.task.title, "Zaktualizowane zadanie")
        self.assertRedirects(
            response, reverse("tasks:detail", kwargs={"pk": self.task.pk})
        )


# ---------------------------------------------------------------------------
# Testy: TaskDeleteView
# ---------------------------------------------------------------------------


class TaskDeleteViewTest(TestCase):
    """Testy widoku usuwania zadania (tylko ADMIN)."""

    def setUp(self) -> None:
        self.admin = _make_user("delete_admin", role=UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("delete_handlowiec")
        self.company = _make_company("Delete Co", owner=self.handlowiec)
        self.task = _make_task(
            "Delete task", assigned_to=self.handlowiec, company=self.company
        )

    def test_admin_can_delete(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("tasks:delete", kwargs={"pk": self.task.pk})
        )
        self.assertFalse(Task.objects.filter(pk=self.task.pk).exists())
        self.assertRedirects(response, reverse("tasks:list"))

    def test_handlowiec_gets_403_on_delete(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.post(
            reverse("tasks:delete", kwargs={"pk": self.task.pk})
        )
        self.assertEqual(response.status_code, 403)
        self.assertTrue(Task.objects.filter(pk=self.task.pk).exists())

    def test_delete_confirm_page_returns_200_for_admin(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:delete", kwargs={"pk": self.task.pk}))
        self.assertEqual(response.status_code, 200)


# ---------------------------------------------------------------------------
# Testy: TaskCompleteView i TaskCancelView
# ---------------------------------------------------------------------------


class TaskCompleteAndCancelViewTest(TestCase):
    """Testy akcji zmiany statusu zadania."""

    def setUp(self) -> None:
        self.owner = _make_user("action_owner")
        self.other = _make_user("action_other")
        self.company = _make_company("Action Co", owner=self.owner)

    def _fresh_task(self, status: str = Task.Status.DO_ZROBIENIA) -> Task:
        return _make_task(
            "Akcja task", assigned_to=self.owner, company=self.company, status=status
        )

    def test_complete_sets_status_wykonane(self) -> None:
        task = self._fresh_task()
        self.client.force_login(self.owner)
        self.client.post(reverse("tasks:complete", kwargs={"pk": task.pk}))
        task.refresh_from_db()
        self.assertEqual(task.status, Task.Status.WYKONANE)

    def test_complete_redirects_to_detail(self) -> None:
        task = self._fresh_task()
        self.client.force_login(self.owner)
        response = self.client.post(reverse("tasks:complete", kwargs={"pk": task.pk}))
        self.assertRedirects(response, reverse("tasks:detail", kwargs={"pk": task.pk}))

    def test_cancel_sets_status_anulowane(self) -> None:
        task = self._fresh_task()
        self.client.force_login(self.owner)
        self.client.post(reverse("tasks:cancel", kwargs={"pk": task.pk}))
        task.refresh_from_db()
        self.assertEqual(task.status, Task.Status.ANULOWANE)

    def test_cancel_redirects_to_detail(self) -> None:
        task = self._fresh_task()
        self.client.force_login(self.owner)
        response = self.client.post(reverse("tasks:cancel", kwargs={"pk": task.pk}))
        self.assertRedirects(response, reverse("tasks:detail", kwargs={"pk": task.pk}))

    def test_other_handlowiec_cannot_complete(self) -> None:
        task = self._fresh_task()
        self.client.force_login(self.other)
        response = self.client.post(reverse("tasks:complete", kwargs={"pk": task.pk}))
        self.assertEqual(response.status_code, 403)
        task.refresh_from_db()
        self.assertEqual(task.status, Task.Status.DO_ZROBIENIA)

    def test_other_handlowiec_cannot_cancel(self) -> None:
        task = self._fresh_task()
        self.client.force_login(self.other)
        response = self.client.post(reverse("tasks:cancel", kwargs={"pk": task.pk}))
        self.assertEqual(response.status_code, 403)

    def test_complete_cancelled_task_shows_error_message(self) -> None:
        task = self._fresh_task(status=Task.Status.ANULOWANE)
        self.client.force_login(self.owner)
        response = self.client.post(reverse("tasks:complete", kwargs={"pk": task.pk}))
        # Powinien przekierowac z komunikatem bledu, nie wyjatkiem 500
        self.assertEqual(response.status_code, 302)
        task.refresh_from_db()
        self.assertEqual(task.status, Task.Status.ANULOWANE)

    def test_cancel_completed_task_shows_error_message(self) -> None:
        task = self._fresh_task(status=Task.Status.WYKONANE)
        self.client.force_login(self.owner)
        response = self.client.post(reverse("tasks:cancel", kwargs={"pk": task.pk}))
        self.assertEqual(response.status_code, 302)
        task.refresh_from_db()
        self.assertEqual(task.status, Task.Status.WYKONANE)


# ---------------------------------------------------------------------------
# TaskExportView
# ---------------------------------------------------------------------------


class TaskExportViewTest(TestCase):
    """Testy eksportu zadan do XLSX."""

    def setUp(self) -> None:
        self.admin = _make_user("texp_admin", role=UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("texp_hand")
        self.other = _make_user("texp_other")
        self.company = _make_company("Exp Co", owner=self.handlowiec)
        self.own_task = _make_task(
            "Moje Zadanie", assigned_to=self.handlowiec, company=self.company
        )
        self.other_task = _make_task(
            "Cudze Zadanie", assigned_to=self.other, company=self.company
        )

    def test_export_redirect_anonymous(self) -> None:
        """Anonimowy uzytkownik jest przekierowywany do logowania."""
        response = self.client.get(reverse("tasks:export_xlsx"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_export_returns_200_for_logged_user(self) -> None:
        """Zalogowany uzytkownik otrzymuje odpowiedz 200."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:export_xlsx"))
        self.assertEqual(response.status_code, 200)

    def test_export_content_type(self) -> None:
        """Odpowiedz ma Content-Type xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:export_xlsx"))
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])

    def test_export_content_disposition_attachment(self) -> None:
        """Odpowiedz zwraca zalacznik z rozszerzeniem .xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:export_xlsx"))
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_handlowiec_exports_only_own_tasks(self) -> None:
        """HANDLOWIEC eksportuje tylko swoje zadania."""
        import io as _io

        import openpyxl

        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("tasks:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moje Zadanie", titles)
        self.assertNotIn("Cudze Zadanie", titles)

    def test_admin_exports_all_tasks(self) -> None:
        """ADMIN eksportuje wszystkie zadania."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moje Zadanie", titles)
        self.assertIn("Cudze Zadanie", titles)

    def test_export_has_header_row(self) -> None:
        """Plik XLSX zawiera wiersz naglowkowy z polem ID i Tytul."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("tasks:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Tytuł")
