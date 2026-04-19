"""Testy widokow CRUD aplikacji companies."""

from __future__ import annotations

import json
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import UserProfile
from apps.companies.models import Company

# ---------------------------------------------------------------------------
# Pomocnicze funkcje
# ---------------------------------------------------------------------------


def _make_user(username: str, role: str = UserProfile.Role.HANDLOWIEC) -> User:
    """Tworzy uzytkownika z profilem o podanej roli."""
    user = User.objects.create_user(username=username, password="testpass")
    user.profile.role = role
    user.profile.save()
    return user


def _make_company(name: str, owner: User | None = None) -> Company:
    """Tworzy firme testowa."""
    return Company.objects.create(name=name, owner=owner)


# ---------------------------------------------------------------------------
# Testy: przekierowanie niezalogowanych
# ---------------------------------------------------------------------------


class CompanyViewsAuthRedirectTest(TestCase):
    """Niezalogowani uzytkowniczy musza byc przekierowani na login."""

    def setUp(self) -> None:
        self.company = _make_company("Acme SA")

    def _assert_redirects_to_login(self, url: str, method: str = "get") -> None:
        response = getattr(self.client, method)(url)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_list_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("companies:list"))

    def test_detail_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )

    def test_create_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("companies:create"))

    def test_update_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("companies:update", kwargs={"pk": self.company.pk})
        )

    def test_delete_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("companies:delete", kwargs={"pk": self.company.pk})
        )


# ---------------------------------------------------------------------------
# Testy: CompanyListView
# ---------------------------------------------------------------------------


class CompanyListViewTest(TestCase):
    """Testy widoku listy firm."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec1")
        self.other = _make_user("handlowiec2")
        self.own_company = _make_company("Moja Firma", owner=self.handlowiec)
        self.other_company = _make_company("Cudznia Firma", owner=self.other)

    def test_handlowiec_sees_only_own_companies(self) -> None:
        """HANDLOWIEC widzi tylko swoje firmy."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:list"))
        self.assertEqual(response.status_code, 200)
        companies = list(response.context["companies"])
        self.assertIn(self.own_company, companies)
        self.assertNotIn(self.other_company, companies)

    def test_admin_sees_all_companies(self) -> None:
        """ADMIN widzi wszystkie firmy."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("companies:list"))
        self.assertEqual(response.status_code, 200)
        companies = list(response.context["companies"])
        self.assertIn(self.own_company, companies)
        self.assertIn(self.other_company, companies)

    def test_filter_by_name(self) -> None:
        """Filtr name zaweza wyniki do pasujacych nazw."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("companies:list"), {"name": "Moja"})
        companies = list(response.context["companies"])
        self.assertIn(self.own_company, companies)
        self.assertNotIn(self.other_company, companies)

    def test_filter_by_city(self) -> None:
        """Filtr city zaweza wyniki do podanego miasta."""
        Company.objects.create(name="Warszawa Corp", city="Warszawa")
        self.client.force_login(self.admin)
        response = self.client.get(reverse("companies:list"), {"city": "Warszawa"})
        names = [c.name for c in response.context["companies"]]
        self.assertIn("Warszawa Corp", names)
        self.assertNotIn("Moja Firma", names)

    def test_filter_by_industry(self) -> None:
        """Filtr industry zaweza wyniki do podanej branzy."""
        Company.objects.create(name="Tech Co", industry=Company.Industry.IT)
        self.client.force_login(self.admin)
        response = self.client.get(reverse("companies:list"), {"industry": "IT"})
        names = [c.name for c in response.context["companies"]]
        self.assertIn("Tech Co", names)

    def test_uses_correct_template(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:list"))
        self.assertTemplateUsed(response, "companies/company_list.html")

    def test_context_contains_industry_choices(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:list"))
        self.assertIn("industry_choices", response.context)


# ---------------------------------------------------------------------------
# Testy: CompanyDetailView
# ---------------------------------------------------------------------------


class CompanyDetailViewTest(TestCase):
    """Testy widoku szczegolowego firmy."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Test SA", owner=self.owner)

    def test_owner_can_view_own_company(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["company"], self.company)

    def test_admin_can_view_any_company(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 200)

    def test_handlowiec_cannot_view_other_company(self) -> None:
        """Handlowiec nie widzi cudzej firmy - dostaje 404."""
        self.client.force_login(self.other)
        response = self.client.get(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 404)

    def test_uses_correct_template(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )
        self.assertTemplateUsed(response, "companies/company_detail.html")

    def test_context_can_edit_true_for_owner(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )
        self.assertTrue(response.context["can_edit"])

    def test_context_can_edit_true_for_admin(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("companies:detail", kwargs={"pk": self.company.pk})
        )
        self.assertTrue(response.context["can_edit"])


# ---------------------------------------------------------------------------
# Testy: CompanyCreateView
# ---------------------------------------------------------------------------


class CompanyCreateViewTest(TestCase):
    """Testy widoku tworzenia firmy."""

    def setUp(self) -> None:
        self.user = _make_user("creator")

    def test_get_returns_200(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("companies:create"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "companies/company_form.html")

    def test_post_creates_company_and_sets_owner(self) -> None:
        """POST z poprawnymi danymi tworzy firme i ustawia owner."""
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("companies:create"),
            data={"name": "Nowa Firma Sp. z o.o.", "industry": "IT", "is_active": True},
        )
        self.assertEqual(Company.objects.count(), 1)
        company = Company.objects.first()
        self.assertEqual(company.owner, self.user)
        self.assertRedirects(
            response,
            reverse("companies:detail", kwargs={"pk": company.pk}),
        )

    def test_post_invalid_data_shows_form_again(self) -> None:
        """POST z brakujaca nazwa zwraca formularz z bledami."""
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("companies:create"),
            data={"name": "", "industry": "IT"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["form"].is_valid())
        self.assertEqual(Company.objects.count(), 0)


# ---------------------------------------------------------------------------
# Testy: CompanyUpdateView
# ---------------------------------------------------------------------------


class CompanyUpdateViewTest(TestCase):
    """Testy widoku edycji firmy."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Stara Nazwa", owner=self.owner)

    def test_owner_can_edit(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(
            reverse("companies:update", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 200)

    def test_admin_can_edit_any_company(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("companies:update", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 200)

    def test_other_handlowiec_gets_403(self) -> None:
        """Inny handlowiec probujacy edytowac cudzą firme dostaje 403."""
        self.client.force_login(self.other)
        response = self.client.get(
            reverse("companies:update", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 403)

    def test_post_updates_company_name(self) -> None:
        """POST z poprawnymi danymi aktualizuje firme."""
        self.client.force_login(self.owner)
        self.client.post(
            reverse("companies:update", kwargs={"pk": self.company.pk}),
            data={"name": "Nowa Nazwa SA", "industry": "IT", "is_active": True},
        )
        self.company.refresh_from_db()
        self.assertEqual(self.company.name, "Nowa Nazwa SA")


# ---------------------------------------------------------------------------
# Testy: CompanyDeleteView
# ---------------------------------------------------------------------------


class CompanyDeleteViewTest(TestCase):
    """Testy widoku usuwania firmy (tylko ADMIN)."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec_user")
        self.company = _make_company("Do Usuniecia", owner=self.admin)

    def test_handlowiec_gets_403_on_delete_get(self) -> None:
        """Handlowiec nie moze wejsc na strone usuwania firmy."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(
            reverse("companies:delete", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_can_view_delete_confirmation(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(
            reverse("companies:delete", kwargs={"pk": self.company.pk})
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "companies/company_confirm_delete.html")

    def test_admin_can_delete_company(self) -> None:
        """ADMIN moze usunac firme przez POST."""
        self.client.force_login(self.admin)
        self.client.post(reverse("companies:delete", kwargs={"pk": self.company.pk}))
        self.assertFalse(Company.objects.filter(pk=self.company.pk).exists())

    def test_delete_redirects_to_list(self) -> None:
        """Po usunieciu nastepuje przekierowanie na liste firm."""
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("companies:delete", kwargs={"pk": self.company.pk})
        )
        self.assertRedirects(response, reverse("companies:list"))

    def test_handlowiec_cannot_delete_via_post(self) -> None:
        """Handlowiec nie moze usunac firmy nawet przez POST."""
        self.client.force_login(self.handlowiec)
        self.client.post(reverse("companies:delete", kwargs={"pk": self.company.pk}))
        # Firma powinna nadal istniesc
        self.assertTrue(Company.objects.filter(pk=self.company.pk).exists())


# ---------------------------------------------------------------------------
# Testy: NipLookupView
# ---------------------------------------------------------------------------

NIP_LOOKUP_URL = reverse("companies:nip_lookup")
_MF_SUBJECT = {
    "name": "ACME Sp. z o.o.",
    "workingAddress": "ul. Testowa 1, 00-001 Warszawa",
}
_MF_RESPONSE = {"result": {"subject": _MF_SUBJECT}}


def _mock_mf_response(
    status_code: int = 200, json_data: dict | None = None
) -> MagicMock:
    mock = MagicMock()
    mock.status_code = status_code
    mock.json.return_value = json_data if json_data is not None else _MF_RESPONSE
    return mock


class NipLookupAuthTest(TestCase):
    """Niezalogowani użytkownicy są przekierowywani."""

    def test_redirect_anonymous(self) -> None:
        response = self.client.get(NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])


def _ajax_get(client, url):
    """Wysyła żądanie GET z nagłówkiem AJAX (wymaga JSON zamiast redirect)."""
    return client.get(url, HTTP_X_REQUESTED_WITH="XMLHttpRequest")


class NipLookupValidationTest(TestCase):
    """Walidacja parametru NIP przed wywołaniem API."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    def test_missing_nip_returns_400(self) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL)
        self.assertEqual(response.status_code, 400)

    def test_too_short_nip_returns_400(self) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=123")
        self.assertEqual(response.status_code, 400)

    def test_non_digit_nip_returns_400(self) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=ABCDEFGHIJ")
        self.assertEqual(response.status_code, 400)

    def test_nip_with_dashes_accepted(self) -> None:
        """NIP z myślnikami (np. 123-456-78-90) powinien przejść walidację."""
        with patch(
            "apps.companies.views.requests.get",
            return_value=_mock_mf_response(),
        ):
            response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=123-456-78-90")
        # 10 cyfr po usunięciu myślników – walidacja przechodzi
        self.assertNotEqual(response.status_code, 400)


class NipLookupSuccessTest(TestCase):
    """Poprawna odpowiedź MF zwraca uzupełnione dane firmy."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    @patch("apps.companies.views.requests.get", return_value=_mock_mf_response())
    def test_returns_200_with_company_data(self, mock_get) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertEqual(data["name"], "ACME Sp. z o.o.")
        self.assertEqual(data["city"], "Warszawa")
        self.assertEqual(data["postal_code"], "00-001")
        self.assertEqual(data["address"], "ul. Testowa 1")

    @patch("apps.companies.views.requests.get")
    def test_calls_mf_api(self, mock_get) -> None:
        """API MF jest wywoływane (z fallbackiem po CEIDG gdy token ustawiony)."""

        def side_effect(url, **kwargs):
            if "ceidg" in url:
                m = MagicMock()
                m.status_code = 404
                return m
            return _mock_mf_response()

        mock_get.side_effect = side_effect
        self.client.get(NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertTrue(mock_get.called)
        # Ostatnie wywołanie to MF
        last_url = mock_get.call_args[0][0]
        self.assertIn("wl-api.mf.gov.pl", last_url)


class NipLookupErrorTest(TestCase):
    """Obsługa błędów: 404, timeout, błąd połączenia."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    @patch(
        "apps.companies.views.requests.get",
        return_value=_mock_mf_response(status_code=404),
    )
    def test_mf_404_returns_404(self, mock_get) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertEqual(response.status_code, 404)
        self.assertIn("error", json.loads(response.content))

    @patch(
        "apps.companies.views.requests.get",
        side_effect=__import__("requests").Timeout,
    )
    def test_timeout_returns_504(self, mock_get) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertEqual(response.status_code, 504)

    @patch(
        "apps.companies.views.requests.get",
        side_effect=__import__("requests").RequestException("conn error"),
    )
    def test_request_exception_returns_502(self, mock_get) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertEqual(response.status_code, 502)

    @patch(
        "apps.companies.views.requests.get",
        return_value=_mock_mf_response(json_data={"result": {"subject": {}}}),
    )
    def test_empty_subject_returns_404(self, mock_get) -> None:
        response = _ajax_get(self.client, NIP_LOOKUP_URL + "?nip=1234567890")
        self.assertEqual(response.status_code, 404)


# ---------------------------------------------------------------------------
# Testy: NipSearchView
# ---------------------------------------------------------------------------

NIP_SEARCH_URL = reverse("companies:nip_search")

_CEIDG_RESPONSE = {
    "firmy": [
        {
            "nazwa": "Testowa JDG Jan Kowalski",
            "adresDzialalnosci": {
                "ulica": "ul. Przykładowa",
                "budynek": "5",
                "miasto": "Warszawa",
                "kod": "00-001",
            },
        }
    ]
}


def _mock_ceidg(status_code: int = 200, json_data: dict | None = None) -> MagicMock:
    mock = MagicMock()
    mock.status_code = status_code
    mock.json.return_value = json_data if json_data is not None else _CEIDG_RESPONSE
    return mock


class NipSearchAuthTest(TestCase):
    """Niezalogowani użytkownicy są przekierowywani."""

    def test_get_redirect_anonymous(self) -> None:
        response = self.client.get(NIP_SEARCH_URL)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_post_redirect_anonymous(self) -> None:
        response = self.client.post(NIP_SEARCH_URL, {"nip": "1234567890"})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])


class NipSearchGetTest(TestCase):
    """GET /companies/nip-search/ zwraca formularz."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    def test_get_returns_200(self) -> None:
        response = self.client.get(NIP_SEARCH_URL)
        self.assertEqual(response.status_code, 200)

    def test_get_contains_form(self) -> None:
        response = self.client.get(NIP_SEARCH_URL)
        self.assertContains(response, 'name="nip"')
        self.assertContains(response, 'method="post"')

    def test_get_passes_next_url(self) -> None:
        response = self.client.get(NIP_SEARCH_URL + "?next=/companies/add/")
        self.assertContains(response, "/companies/add/")


class NipSearchPostValidationTest(TestCase):
    """Walidacja NIP w żądaniu POST."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    def test_post_invalid_nip_renders_form_with_error(self) -> None:
        response = self.client.post(
            NIP_SEARCH_URL, {"nip": "123", "next_url": "/companies/add/"}
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "NIP")

    def test_post_empty_nip_renders_form(self) -> None:
        response = self.client.post(
            NIP_SEARCH_URL, {"nip": "", "next_url": "/companies/add/"}
        )
        self.assertEqual(response.status_code, 200)

    def test_post_non_digit_nip_renders_form(self) -> None:
        response = self.client.post(
            NIP_SEARCH_URL, {"nip": "ABCDEFGHIJ", "next_url": "/companies/add/"}
        )
        self.assertEqual(response.status_code, 200)


class NipSearchPostMfTest(TestCase):
    """POST z prawidłowym NIP — fallback na Białą Listę MF."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    def _post(self, nip: str = "1234567890") -> object:
        return self.client.post(
            NIP_SEARCH_URL, {"nip": nip, "next_url": "/companies/add/"}
        )

    @patch("apps.companies.views.requests.get", return_value=_mock_mf_response())
    def test_valid_nip_redirects_with_data(self, mock_get) -> None:
        response = self._post()
        self.assertEqual(response.status_code, 302)
        location = response["Location"]
        self.assertIn("name=", location)
        self.assertIn("1234567890", location)

    @patch("apps.companies.views.requests.get", return_value=_mock_mf_response())
    def test_redirect_target_is_next_url(self, mock_get) -> None:
        response = self._post()
        self.assertTrue(response["Location"].startswith("/companies/add/"))

    @patch(
        "apps.companies.views.requests.get",
        return_value=_mock_mf_response(status_code=404),
    )
    def test_mf_404_shows_error_message(self, mock_get) -> None:
        response = self._post()
        self.assertEqual(response.status_code, 200)
        messages = list(response.wsgi_request._messages)
        self.assertTrue(any("znaleziono" in str(m).lower() for m in messages))


class NipSearchPostCeidgTest(TestCase):
    """POST z prawidłowym NIP — CEIDG jako źródło danych."""

    def setUp(self) -> None:
        self.user = User.objects.create_user("tester", password="pass")
        self.client.force_login(self.user)

    @patch("apps.companies.views.requests.get", return_value=_mock_ceidg())
    def test_ceidg_data_in_redirect(self, mock_get) -> None:
        response = self.client.post(
            NIP_SEARCH_URL, {"nip": "1234567890", "next_url": "/companies/add/"}
        )
        self.assertEqual(response.status_code, 302)
        location = response["Location"]
        self.assertIn("Testowa", location)
        self.assertIn("Warszawa", location)

    @patch("apps.companies.views.requests.get")
    def test_ceidg_404_falls_back_to_mf(self, mock_get) -> None:
        """Gdy CEIDG zwraca 404, widok odpytuje MF jako fallback."""

        def side_effect(url, **kwargs):
            if "ceidg" in url:
                return _mock_ceidg(status_code=404)
            return _mock_mf_response()

        mock_get.side_effect = side_effect
        response = self.client.post(
            NIP_SEARCH_URL, {"nip": "1234567890", "next_url": "/companies/add/"}
        )
        self.assertEqual(response.status_code, 302)
        self.assertGreaterEqual(mock_get.call_count, 2)


# ---------------------------------------------------------------------------
# CompanyExportView
# ---------------------------------------------------------------------------


class CompanyExportViewTest(TestCase):
    """Testy eksportu firm do XLSX."""

    def setUp(self) -> None:
        self.admin = _make_user("exp_admin", role=UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("exp_hand")
        self.other = _make_user("exp_other")
        self.own = _make_company("Moja Firma", owner=self.handlowiec)
        self.other_co = _make_company("Cudza Firma", owner=self.other)

    def test_export_redirect_anonymous(self) -> None:
        """Anonimowy uzytkownik jest przekierowywany do logowania."""
        response = self.client.get(reverse("companies:export_xlsx"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_export_returns_200_for_logged_user(self) -> None:
        """Zalogowany uzytkownik otrzymuje odpowiedz 200."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:export_xlsx"))
        self.assertEqual(response.status_code, 200)

    def test_export_content_type(self) -> None:
        """Odpowiedz ma Content-Type xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:export_xlsx"))
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])

    def test_export_content_disposition_attachment(self) -> None:
        """Odpowiedz zwraca zalacznik z rozszerzeniem .xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:export_xlsx"))
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_handlowiec_exports_only_own_companies(self) -> None:
        """HANDLOWIEC eksportuje tylko swoje firmy (weryfikacja przez zawartosc XLSX)."""
        import io as _io

        import openpyxl

        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("companies:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        names = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moja Firma", names)
        self.assertNotIn("Cudza Firma", names)

    def test_admin_exports_all_companies(self) -> None:
        """ADMIN eksportuje wszystkie firmy."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("companies:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        names = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moja Firma", names)
        self.assertIn("Cudza Firma", names)

    def test_export_has_header_row(self) -> None:
        """Plik XLSX zawiera wiersz naglowkowy."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("companies:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Nazwa")
