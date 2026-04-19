"""Testy widokow CRUD aplikacji deals."""

from __future__ import annotations

import datetime

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import UserProfile
from apps.companies.models import Company
from apps.deals.models import Deal
from apps.leads.models import Lead, WorkflowStage

# ---------------------------------------------------------------------------
# Pomocnicze funkcje
# ---------------------------------------------------------------------------


def _make_user(username: str, role: str = UserProfile.Role.HANDLOWIEC) -> User:
    """Tworzy uzytkownika z profilem o podanej roli."""
    user = User.objects.create_user(username=username, password="testpass")
    user.profile.role = role
    user.profile.save()
    return user


def _make_company(name: str, owner: User | None = None) -> Company:
    """Tworzy firme testowa."""
    return Company.objects.create(name=name, owner=owner)


def _make_deal(
    title: str,
    company: Company,
    owner: User | None = None,
    status: str = Deal.Status.AKTYWNA,
    lead: Lead | None = None,
) -> Deal:
    """Tworzy umowe testowa."""
    return Deal.objects.create(
        title=title,
        company=company,
        owner=owner,
        status=status,
        lead=lead,
        value="10000.00",
        close_date=datetime.date.today() + datetime.timedelta(days=30),
    )


def _make_lead(title: str, company: Company, owner: User | None = None) -> Lead:
    """Tworzy lead testowy."""
    stage, _ = WorkflowStage.objects.get_or_create(
        order=1, defaults={"name": "Nowy", "color": "#6c757d"}
    )
    return Lead.objects.create(title=title, company=company, owner=owner, stage=stage)


# ---------------------------------------------------------------------------
# Testy: przekierowanie niezalogowanych
# ---------------------------------------------------------------------------


class DealViewsAuthRedirectTest(TestCase):
    """Niezalogowani uzytkowniczy musza byc przekierowani na login."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.company = _make_company("Acme SA", owner=self.owner)
        self.deal = _make_deal("Test Deal", self.company, owner=self.owner)

    def _assert_redirects_to_login(self, url: str, method: str = "get") -> None:
        response = getattr(self.client, method)(url)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_list_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("deals:list"))

    def test_detail_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("deals:detail", kwargs={"pk": self.deal.pk})
        )

    def test_create_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("deals:create"))

    def test_update_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("deals:update", kwargs={"pk": self.deal.pk})
        )

    def test_delete_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("deals:delete", kwargs={"pk": self.deal.pk})
        )

    def test_complete_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("deals:complete", kwargs={"pk": self.deal.pk}), method="post"
        )

    def test_cancel_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("deals:cancel", kwargs={"pk": self.deal.pk}), method="post"
        )


# ---------------------------------------------------------------------------
# Testy: DealListView
# ---------------------------------------------------------------------------


class DealListViewTest(TestCase):
    """Testy widoku listy umow."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec1")
        self.other = _make_user("handlowiec2")
        self.company = _make_company("Firma A", owner=self.handlowiec)
        self.other_company = _make_company("Firma B", owner=self.other)
        self.own_deal = _make_deal("Moja Umowa", self.company, owner=self.handlowiec)
        self.other_deal = _make_deal(
            "Cudza Umowa", self.other_company, owner=self.other
        )

    def test_handlowiec_sees_only_own_deals(self) -> None:
        """HANDLOWIEC widzi tylko swoje umowy."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:list"))
        self.assertEqual(response.status_code, 200)
        deals = list(response.context["deals"])
        self.assertIn(self.own_deal, deals)
        self.assertNotIn(self.other_deal, deals)

    def test_admin_sees_all_deals(self) -> None:
        """ADMIN widzi wszystkie umowy."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:list"))
        deals = list(response.context["deals"])
        self.assertIn(self.own_deal, deals)
        self.assertIn(self.other_deal, deals)

    def test_filter_by_status(self) -> None:
        """Filtr status zaweza wyniki."""
        zrealizowana = _make_deal(
            "Stara Umowa",
            self.company,
            owner=self.admin,
            status=Deal.Status.ZREALIZOWANA,
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:list"), {"status": "ZREALIZOWANA"})
        deals = list(response.context["deals"])
        self.assertIn(zrealizowana, deals)
        self.assertNotIn(self.own_deal, deals)

    def test_filter_by_company(self) -> None:
        """Filtr company zaweza wyniki do podanej firmy."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:list"), {"company": "Firma A"})
        deals = list(response.context["deals"])
        self.assertIn(self.own_deal, deals)
        self.assertNotIn(self.other_deal, deals)

    def test_context_contains_active_total(self) -> None:
        """Kontekst zawiera sume wartosci aktywnych umow."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:list"))
        self.assertIn("active_total", response.context)
        self.assertGreaterEqual(response.context["active_total"], 0)

    def test_uses_correct_template(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:list"))
        self.assertTemplateUsed(response, "deals/deal_list.html")

    def test_context_contains_status_choices(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:list"))
        self.assertIn("status_choices", response.context)


# ---------------------------------------------------------------------------
# Testy: DealDetailView
# ---------------------------------------------------------------------------


class DealDetailViewTest(TestCase):
    """Testy widoku szczegolowego umowy."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Test SA", owner=self.owner)
        self.deal = _make_deal("Test Deal", self.company, owner=self.owner)

    def test_owner_can_view_own_deal(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("deals:detail", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["deal"], self.deal)

    def test_admin_can_view_any_deal(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:detail", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 200)

    def test_handlowiec_cannot_view_other_deal(self) -> None:
        """Handlowiec nie widzi cudzej umowy — dostaje 404."""
        self.client.force_login(self.other)
        response = self.client.get(reverse("deals:detail", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 404)

    def test_uses_correct_template(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("deals:detail", kwargs={"pk": self.deal.pk}))
        self.assertTemplateUsed(response, "deals/deal_detail.html")

    def test_context_can_edit_true_for_owner(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("deals:detail", kwargs={"pk": self.deal.pk}))
        self.assertTrue(response.context["can_edit"])

    def test_context_can_edit_true_for_admin(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:detail", kwargs={"pk": self.deal.pk}))
        self.assertTrue(response.context["can_edit"])


# ---------------------------------------------------------------------------
# Testy: DealCreateView
# ---------------------------------------------------------------------------


class DealCreateViewTest(TestCase):
    """Testy widoku tworzenia umowy."""

    def setUp(self) -> None:
        self.user = _make_user("creator")
        self.company = _make_company("Moja Firma", owner=self.user)

    def test_get_returns_200(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("deals:create"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "deals/deal_form.html")

    def test_post_creates_deal_and_sets_owner(self) -> None:
        """POST z poprawnymi danymi tworzy umowe i ustawia owner."""
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("deals:create"),
            data={
                "title": "Nowa Umowa Testowa",
                "company": self.company.pk,
                "value": "25000.00",
                "close_date": "2026-12-31",
                "description": "",
            },
        )
        self.assertEqual(Deal.objects.count(), 1)
        deal = Deal.objects.first()
        self.assertEqual(deal.owner, self.user)
        self.assertRedirects(response, reverse("deals:detail", kwargs={"pk": deal.pk}))

    def test_post_invalid_data_shows_form_again(self) -> None:
        """POST z brakujacym tytulem zwraca formularz z bledami."""
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("deals:create"),
            data={"title": "", "company": self.company.pk, "close_date": "2026-12-31"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["form"].is_valid())
        self.assertEqual(Deal.objects.count(), 0)

    def test_get_with_lead_id_prefills_lead_and_company(self) -> None:
        """GET z ?lead_id= wstepnie wypelnia pola lead i company."""
        lead = _make_lead("Testowy Lead", self.company, owner=self.user)
        self.client.force_login(self.user)
        response = self.client.get(reverse("deals:create"), {"lead_id": lead.pk})
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertEqual(form.initial.get("lead"), lead)
        self.assertEqual(form.initial.get("company"), self.company)

    def test_handlowiec_company_queryset_limited_to_own(self) -> None:
        """Formularz dla HANDLOWCA zawiera tylko jego firmy."""
        other = _make_user("other_user")
        _make_company("Cudza Firma", owner=other)
        self.client.force_login(self.user)
        response = self.client.get(reverse("deals:create"))
        form = response.context["form"]
        companies = list(form.fields["company"].queryset)
        self.assertIn(self.company, companies)
        for comp in companies:
            self.assertNotEqual(comp.name, "Cudza Firma")

    def test_admin_company_queryset_contains_all(self) -> None:
        """Formularz dla ADMINA zawiera wszystkie firmy."""
        admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        other = _make_user("other_user")
        other_company = _make_company("Cudza Firma", owner=other)
        self.client.force_login(admin)
        response = self.client.get(reverse("deals:create"))
        form = response.context["form"]
        companies = list(form.fields["company"].queryset)
        self.assertIn(self.company, companies)
        self.assertIn(other_company, companies)


# ---------------------------------------------------------------------------
# Testy: DealUpdateView
# ---------------------------------------------------------------------------


class DealUpdateViewTest(TestCase):
    """Testy widoku edycji umowy."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Firma", owner=self.owner)
        self.deal = _make_deal("Stara Umowa", self.company, owner=self.owner)

    def test_owner_can_edit(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("deals:update", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 200)

    def test_admin_can_edit_any_deal(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:update", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 200)

    def test_other_handlowiec_gets_403(self) -> None:
        """Inny handlowiec probujacy edytowac cudza umowe dostaje 403."""
        self.client.force_login(self.other)
        response = self.client.get(reverse("deals:update", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 403)

    def test_post_updates_deal_title(self) -> None:
        """POST z poprawnymi danymi aktualizuje umowe."""
        self.client.force_login(self.owner)
        self.client.post(
            reverse("deals:update", kwargs={"pk": self.deal.pk}),
            data={
                "title": "Zaktualizowana Umowa",
                "company": self.company.pk,
                "value": "15000.00",
                "close_date": "2026-12-31",
                "description": "",
            },
        )
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.title, "Zaktualizowana Umowa")


# ---------------------------------------------------------------------------
# Testy: DealDeleteView
# ---------------------------------------------------------------------------


class DealDeleteViewTest(TestCase):
    """Testy widoku usuwania umowy (tylko ADMIN)."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec_user")
        self.company = _make_company("Firma", owner=self.admin)
        self.deal = _make_deal("Do Usuniecia", self.company, owner=self.admin)

    def test_handlowiec_gets_403_on_delete_get(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:delete", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 403)

    def test_admin_can_view_delete_confirmation(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:delete", kwargs={"pk": self.deal.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "deals/deal_confirm_delete.html")

    def test_admin_can_delete_deal(self) -> None:
        self.client.force_login(self.admin)
        self.client.post(reverse("deals:delete", kwargs={"pk": self.deal.pk}))
        self.assertFalse(Deal.objects.filter(pk=self.deal.pk).exists())

    def test_delete_redirects_to_list(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("deals:delete", kwargs={"pk": self.deal.pk})
        )
        self.assertRedirects(response, reverse("deals:list"))

    def test_handlowiec_cannot_delete_via_post(self) -> None:
        self.client.force_login(self.handlowiec)
        self.client.post(reverse("deals:delete", kwargs={"pk": self.deal.pk}))
        self.assertTrue(Deal.objects.filter(pk=self.deal.pk).exists())


# ---------------------------------------------------------------------------
# Testy: DealCompleteView i DealCancelView
# ---------------------------------------------------------------------------


class DealCompleteViewTest(TestCase):
    """Testy widoku zatwierdzania umowy jako zrealizowanej."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Firma", owner=self.owner)
        self.deal = _make_deal("Aktywna Umowa", self.company, owner=self.owner)

    def test_owner_can_complete_deal(self) -> None:
        """Wlasciciel moze zatwierdzic umowe jako zrealizowana."""
        self.client.force_login(self.owner)
        self.client.post(reverse("deals:complete", kwargs={"pk": self.deal.pk}))
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.ZREALIZOWANA)
        self.assertIsNotNone(self.deal.signed_at)

    def test_admin_can_complete_any_deal(self) -> None:
        """ADMIN moze zatwierdzic dowolna umowe."""
        self.client.force_login(self.admin)
        self.client.post(reverse("deals:complete", kwargs={"pk": self.deal.pk}))
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.ZREALIZOWANA)

    def test_other_handlowiec_gets_403_on_complete(self) -> None:
        """Inny handlowiec nie moze zatwierdzic cudzej umowy."""
        self.client.force_login(self.other)
        response = self.client.post(
            reverse("deals:complete", kwargs={"pk": self.deal.pk})
        )
        self.assertEqual(response.status_code, 403)
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.AKTYWNA)

    def test_complete_redirects_to_detail(self) -> None:
        """Po zatwierdzeniu nastepuje przekierowanie na szczegoly."""
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse("deals:complete", kwargs={"pk": self.deal.pk})
        )
        self.assertRedirects(
            response, reverse("deals:detail", kwargs={"pk": self.deal.pk})
        )

    def test_cannot_complete_cancelled_deal(self) -> None:
        """Nie mozna zatwierdzic anulowanej umowy."""
        self.deal.status = Deal.Status.ANULOWANA
        self.deal.save()
        self.client.force_login(self.owner)
        self.client.post(reverse("deals:complete", kwargs={"pk": self.deal.pk}))
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.ANULOWANA)


class DealCancelViewTest(TestCase):
    """Testy widoku anulowania umowy."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Firma", owner=self.owner)
        self.deal = _make_deal("Aktywna Umowa", self.company, owner=self.owner)

    def test_owner_can_cancel_deal(self) -> None:
        """Wlasciciel moze anulowaç umowe."""
        self.client.force_login(self.owner)
        self.client.post(reverse("deals:cancel", kwargs={"pk": self.deal.pk}))
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.ANULOWANA)

    def test_admin_can_cancel_any_deal(self) -> None:
        """ADMIN moze anulowaç dowolna umowe."""
        self.client.force_login(self.admin)
        self.client.post(reverse("deals:cancel", kwargs={"pk": self.deal.pk}))
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.ANULOWANA)

    def test_other_handlowiec_gets_403_on_cancel(self) -> None:
        """Inny handlowiec nie moze anulowaç cudzej umowy."""
        self.client.force_login(self.other)
        response = self.client.post(
            reverse("deals:cancel", kwargs={"pk": self.deal.pk})
        )
        self.assertEqual(response.status_code, 403)
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.AKTYWNA)

    def test_cancel_redirects_to_detail(self) -> None:
        """Po anulowaniu nastepuje przekierowanie na szczegoly."""
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse("deals:cancel", kwargs={"pk": self.deal.pk})
        )
        self.assertRedirects(
            response, reverse("deals:detail", kwargs={"pk": self.deal.pk})
        )

    def test_cannot_cancel_completed_deal(self) -> None:
        """Nie mozna anulowaç zrealizowanej umowy."""
        self.deal.status = Deal.Status.ZREALIZOWANA
        self.deal.save()
        self.client.force_login(self.owner)
        self.client.post(reverse("deals:cancel", kwargs={"pk": self.deal.pk}))
        self.deal.refresh_from_db()
        self.assertEqual(self.deal.status, Deal.Status.ZREALIZOWANA)


# ---------------------------------------------------------------------------
# DealExportView
# ---------------------------------------------------------------------------


class DealExportViewTest(TestCase):
    """Testy eksportu umow do XLSX."""

    def setUp(self) -> None:
        self.admin = _make_user("dexp_admin", role=UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("dexp_hand")
        self.other = _make_user("dexp_other")
        self.company = _make_company("Exp Co", owner=self.handlowiec)
        self.own_deal = _make_deal(
            "Moja Umowa", company=self.company, owner=self.handlowiec
        )
        self.other_deal = _make_deal(
            "Cudza Umowa", company=self.company, owner=self.other
        )

    def test_export_redirect_anonymous(self) -> None:
        """Anonimowy uzytkownik jest przekierowywany do logowania."""
        response = self.client.get(reverse("deals:export_xlsx"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_export_returns_200_for_logged_user(self) -> None:
        """Zalogowany uzytkownik otrzymuje odpowiedz 200."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:export_xlsx"))
        self.assertEqual(response.status_code, 200)

    def test_export_content_type(self) -> None:
        """Odpowiedz ma Content-Type xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:export_xlsx"))
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])

    def test_export_content_disposition_attachment(self) -> None:
        """Odpowiedz zwraca zalacznik z rozszerzeniem .xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:export_xlsx"))
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_handlowiec_exports_only_own_deals(self) -> None:
        """HANDLOWIEC eksportuje tylko swoje umowy."""
        import io as _io

        import openpyxl

        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("deals:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moja Umowa", titles)
        self.assertNotIn("Cudza Umowa", titles)

    def test_admin_exports_all_deals(self) -> None:
        """ADMIN eksportuje wszystkie umowy."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moja Umowa", titles)
        self.assertIn("Cudza Umowa", titles)

    def test_export_has_header_row(self) -> None:
        """Plik XLSX zawiera wiersz naglowkowy z polem ID i Tytul."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("deals:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Tytuł")
