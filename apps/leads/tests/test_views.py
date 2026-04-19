"""Testy widokow CRUD i Kanban aplikacji leads."""

from __future__ import annotations

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import UserProfile
from apps.companies.models import Company
from apps.leads.models import Lead, WorkflowStage

# ---------------------------------------------------------------------------
# Pomocnicze funkcje
# ---------------------------------------------------------------------------


def _make_user(username: str, role: str = UserProfile.Role.HANDLOWIEC) -> User:
    """Tworzy uzytkownika z profilem o podanej roli."""
    user = User.objects.create_user(username=username, password="testpass")
    user.profile.role = role
    user.profile.save()
    return user


def _make_company(name: str, owner: User | None = None) -> Company:
    """Tworzy firme testowa."""
    return Company.objects.create(name=name, owner=owner)


def _get_stage(order: int = 1) -> WorkflowStage:
    """Zwraca istniejacy etap lub tworzy nowy."""
    stage, _ = WorkflowStage.objects.get_or_create(
        order=order,
        defaults={"name": f"Etap {order}", "color": "#6c757d"},
    )
    return stage


def _make_lead(
    title: str,
    company: Company,
    owner: User | None = None,
    status: str = Lead.Status.NOWY,
    stage: WorkflowStage | None = None,
) -> Lead:
    """Tworzy lead testowy."""
    if stage is None:
        stage = _get_stage()
    return Lead.objects.create(
        title=title,
        company=company,
        owner=owner,
        status=status,
        stage=stage,
    )


# ---------------------------------------------------------------------------
# Testy: przekierowanie niezalogowanych
# ---------------------------------------------------------------------------


class LeadViewsAuthRedirectTest(TestCase):
    """Niezalogowani uzytkowniczy musza byc przekierowani na login."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.company = _make_company("Acme SA", owner=self.owner)
        self.lead = _make_lead("Test Lead", self.company, owner=self.owner)

    def _assert_redirects_to_login(self, url: str, method: str = "get") -> None:
        response = getattr(self.client, method)(url)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_list_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("leads:list"))

    def test_kanban_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("leads:kanban"))

    def test_detail_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("leads:detail", kwargs={"pk": self.lead.pk})
        )

    def test_create_requires_login(self) -> None:
        self._assert_redirects_to_login(reverse("leads:create"))

    def test_update_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("leads:update", kwargs={"pk": self.lead.pk})
        )

    def test_delete_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("leads:delete", kwargs={"pk": self.lead.pk})
        )

    def test_close_requires_login(self) -> None:
        self._assert_redirects_to_login(
            reverse("leads:close", kwargs={"pk": self.lead.pk}), method="post"
        )


# ---------------------------------------------------------------------------
# Testy: LeadListView
# ---------------------------------------------------------------------------


class LeadListViewTest(TestCase):
    """Testy widoku listy leadow."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec1")
        self.other = _make_user("handlowiec2")
        self.company = _make_company("Firma A", owner=self.handlowiec)
        self.other_company = _make_company("Firma B", owner=self.other)
        self.own_lead = _make_lead("Moj Lead", self.company, owner=self.handlowiec)
        self.other_lead = _make_lead("Cudzy Lead", self.other_company, owner=self.other)

    def test_handlowiec_sees_only_own_leads(self) -> None:
        """HANDLOWIEC widzi tylko swoje leady."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:list"))
        self.assertEqual(response.status_code, 200)
        leads = list(response.context["leads"])
        self.assertIn(self.own_lead, leads)
        self.assertNotIn(self.other_lead, leads)

    def test_admin_sees_all_leads(self) -> None:
        """ADMIN widzi wszystkie leady."""
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:list"))
        leads = list(response.context["leads"])
        self.assertIn(self.own_lead, leads)
        self.assertIn(self.other_lead, leads)

    def test_filter_by_status(self) -> None:
        """Filtr status zaweza wyniki."""
        closed = _make_lead(
            "Wygrana",
            self.company,
            owner=self.admin,
            status=Lead.Status.WYGRANA,
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:list"), {"status": "WYGRANA"})
        leads = list(response.context["leads"])
        self.assertIn(closed, leads)
        self.assertNotIn(self.own_lead, leads)

    def test_filter_by_source(self) -> None:
        """Filtr source zaweza wyniki."""
        stage = _get_stage()
        lead_polecenie = Lead.objects.create(
            title="Polecenie",
            company=self.company,
            owner=self.admin,
            source=Lead.Source.POLECENIE,
            stage=stage,
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:list"), {"source": "POLECENIE"})
        leads = list(response.context["leads"])
        self.assertIn(lead_polecenie, leads)

    def test_filter_by_stage(self) -> None:
        """Filtr stage zaweza wyniki do wybranego etapu."""
        stage2, _ = WorkflowStage.objects.get_or_create(
            name="Oferta", defaults={"order": 3, "color": "#fd7e14"}
        )
        Lead.objects.create(
            title="Lead Oferta",
            company=self.company,
            owner=self.admin,
            stage=stage2,
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:list"), {"stage": str(stage2.pk)})
        leads = list(response.context["leads"])
        titles = [lead.title for lead in leads]
        self.assertIn("Lead Oferta", titles)

    def test_uses_correct_template(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:list"))
        self.assertTemplateUsed(response, "leads/lead_list.html")

    def test_context_contains_status_choices(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:list"))
        self.assertIn("status_choices", response.context)


# ---------------------------------------------------------------------------
# Testy: LeadKanbanView
# ---------------------------------------------------------------------------


class LeadKanbanViewTest(TestCase):
    """Testy widoku Kanban."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec1")
        self.other = _make_user("handlowiec2")
        self.company = _make_company("Firma A", owner=self.handlowiec)
        self.stage = _get_stage()
        self.own_lead = _make_lead(
            "Moj Lead", self.company, owner=self.handlowiec, stage=self.stage
        )

    def test_kanban_returns_200(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:kanban"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "leads/lead_kanban.html")

    def test_kanban_context_has_columns(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:kanban"))
        self.assertIn("kanban_columns", response.context)

    def test_handlowiec_kanban_shows_only_own_leads(self) -> None:
        """HANDLOWIEC widzi na Kanbanie tylko swoje leady."""
        other_company = _make_company("Firma B", owner=self.other)
        other_lead = _make_lead(
            "Cudzy Lead", other_company, owner=self.other, stage=self.stage
        )
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:kanban"))
        all_leads = []
        for col in response.context["kanban_columns"]:
            all_leads.extend(col["leads"])
        self.assertIn(self.own_lead, all_leads)
        self.assertNotIn(other_lead, all_leads)

    def test_kanban_excludes_closed_leads(self) -> None:
        """Kanban nie pokazuje zamknietych leadow."""
        closed = _make_lead(
            "Zamkniety",
            self.company,
            owner=self.handlowiec,
            status=Lead.Status.WYGRANA,
            stage=self.stage,
        )
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:kanban"))
        all_leads = []
        for col in response.context["kanban_columns"]:
            all_leads.extend(col["leads"])
        self.assertNotIn(closed, all_leads)
        self.assertIn(self.own_lead, all_leads)


# ---------------------------------------------------------------------------
# Testy: LeadDetailView
# ---------------------------------------------------------------------------


class LeadDetailViewTest(TestCase):
    """Testy widoku szczegolowego leada."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Test SA", owner=self.owner)
        self.lead = _make_lead("Test Lead", self.company, owner=self.owner)

    def test_owner_can_view_own_lead(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("leads:detail", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["lead"], self.lead)

    def test_admin_can_view_any_lead(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:detail", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 200)

    def test_handlowiec_cannot_view_other_lead(self) -> None:
        """Handlowiec nie widzi cudzego leada — dostaje 404."""
        self.client.force_login(self.other)
        response = self.client.get(reverse("leads:detail", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 404)

    def test_uses_correct_template(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("leads:detail", kwargs={"pk": self.lead.pk}))
        self.assertTemplateUsed(response, "leads/lead_detail.html")

    def test_context_can_edit_true_for_owner(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("leads:detail", kwargs={"pk": self.lead.pk}))
        self.assertTrue(response.context["can_edit"])

    def test_context_can_edit_true_for_admin(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:detail", kwargs={"pk": self.lead.pk}))
        self.assertTrue(response.context["can_edit"])


# ---------------------------------------------------------------------------
# Testy: LeadCreateView
# ---------------------------------------------------------------------------


class LeadCreateViewTest(TestCase):
    """Testy widoku tworzenia leada."""

    def setUp(self) -> None:
        self.user = _make_user("creator")
        self.company = _make_company("Moja Firma", owner=self.user)
        self.stage = _get_stage()

    def test_get_returns_200(self) -> None:
        self.client.force_login(self.user)
        response = self.client.get(reverse("leads:create"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "leads/lead_form.html")

    def test_post_creates_lead_and_sets_owner(self) -> None:
        """POST z poprawnymi danymi tworzy lead i ustawia owner."""
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("leads:create"),
            data={
                "title": "Nowy Lead Testowy",
                "company": self.company.pk,
                "source": Lead.Source.POLECENIE,
                "value": "5000.00",
                "stage": self.stage.pk,
                "description": "",
            },
        )
        self.assertEqual(Lead.objects.count(), 1)
        lead = Lead.objects.first()
        self.assertEqual(lead.owner, self.user)
        self.assertRedirects(response, reverse("leads:detail", kwargs={"pk": lead.pk}))

    def test_post_invalid_data_shows_form_again(self) -> None:
        """POST z brakujacym tytulem zwraca formularz z bledami."""
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("leads:create"),
            data={"title": "", "company": self.company.pk, "stage": self.stage.pk},
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["form"].is_valid())
        self.assertEqual(Lead.objects.count(), 0)

    def test_handlowiec_company_queryset_limited_to_own(self) -> None:
        """Formularz dla HANDLOWCA zawiera tylko jego firmy."""
        other = _make_user("other_user")
        _make_company("Cudza Firma", owner=other)
        self.client.force_login(self.user)
        response = self.client.get(reverse("leads:create"))
        form = response.context["form"]
        companies = list(form.fields["company"].queryset)
        self.assertIn(self.company, companies)
        for c in companies:
            self.assertNotEqual(c.name, "Cudza Firma")


# ---------------------------------------------------------------------------
# Testy: LeadUpdateView
# ---------------------------------------------------------------------------


class LeadUpdateViewTest(TestCase):
    """Testy widoku edycji leada."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Firma", owner=self.owner)
        self.stage = _get_stage()
        self.lead = _make_lead("Stary Tytul", self.company, owner=self.owner)

    def test_owner_can_edit(self) -> None:
        self.client.force_login(self.owner)
        response = self.client.get(reverse("leads:update", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 200)

    def test_admin_can_edit_any_lead(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:update", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 200)

    def test_other_handlowiec_gets_403(self) -> None:
        """Inny handlowiec probujacy edytowac cudzy lead dostaje 403."""
        self.client.force_login(self.other)
        response = self.client.get(reverse("leads:update", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 403)

    def test_post_updates_lead_title(self) -> None:
        """POST z poprawnymi danymi aktualizuje lead."""
        self.client.force_login(self.owner)
        self.client.post(
            reverse("leads:update", kwargs={"pk": self.lead.pk}),
            data={
                "title": "Nowy Tytul",
                "company": self.company.pk,
                "source": Lead.Source.INNE,
                "value": "1000.00",
                "stage": self.stage.pk,
                "description": "",
            },
        )
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.title, "Nowy Tytul")


# ---------------------------------------------------------------------------
# Testy: LeadDeleteView
# ---------------------------------------------------------------------------


class LeadDeleteViewTest(TestCase):
    """Testy widoku usuwania leada (tylko ADMIN)."""

    def setUp(self) -> None:
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("handlowiec_user")
        self.company = _make_company("Firma", owner=self.admin)
        self.lead = _make_lead("Do Usuniecia", self.company, owner=self.admin)

    def test_handlowiec_gets_403_on_delete_get(self) -> None:
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:delete", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 403)

    def test_admin_can_view_delete_confirmation(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:delete", kwargs={"pk": self.lead.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "leads/lead_confirm_delete.html")

    def test_admin_can_delete_lead(self) -> None:
        self.client.force_login(self.admin)
        self.client.post(reverse("leads:delete", kwargs={"pk": self.lead.pk}))
        self.assertFalse(Lead.objects.filter(pk=self.lead.pk).exists())

    def test_delete_redirects_to_list(self) -> None:
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("leads:delete", kwargs={"pk": self.lead.pk})
        )
        self.assertRedirects(response, reverse("leads:list"))

    def test_handlowiec_cannot_delete_via_post(self) -> None:
        self.client.force_login(self.handlowiec)
        self.client.post(reverse("leads:delete", kwargs={"pk": self.lead.pk}))
        self.assertTrue(Lead.objects.filter(pk=self.lead.pk).exists())


# ---------------------------------------------------------------------------
# Testy: LeadCloseView
# ---------------------------------------------------------------------------


class LeadCloseViewTest(TestCase):
    """Testy widoku zamkniecia leada."""

    def setUp(self) -> None:
        self.owner = _make_user("owner_user")
        self.other = _make_user("other_user")
        self.admin = _make_user("admin_user", UserProfile.Role.ADMIN)
        self.company = _make_company("Firma", owner=self.owner)
        self.lead = _make_lead("Aktywny Lead", self.company, owner=self.owner)

    def test_owner_can_close_lead_as_wygrana(self) -> None:
        """Wlasciciel moze zamknac lead jako Wygrana."""
        self.client.force_login(self.owner)
        self.client.post(
            reverse("leads:close", kwargs={"pk": self.lead.pk}),
            data={"close_status": "WYGRANA"},
        )
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.status, Lead.Status.WYGRANA)
        self.assertIsNotNone(self.lead.closed_at)

    def test_owner_can_close_lead_as_przegrana(self) -> None:
        """Wlasciciel moze zamknac lead jako Przegrana."""
        self.client.force_login(self.owner)
        self.client.post(
            reverse("leads:close", kwargs={"pk": self.lead.pk}),
            data={"close_status": "PRZEGRANA"},
        )
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.status, Lead.Status.PRZEGRANA)

    def test_admin_can_close_any_lead(self) -> None:
        """ADMIN moze zamknac dowolny lead."""
        self.client.force_login(self.admin)
        self.client.post(
            reverse("leads:close", kwargs={"pk": self.lead.pk}),
            data={"close_status": "WYGRANA"},
        )
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.status, Lead.Status.WYGRANA)

    def test_other_handlowiec_gets_403(self) -> None:
        """Inny handlowiec nie moze zamknac cudzego leada."""
        self.client.force_login(self.other)
        response = self.client.post(
            reverse("leads:close", kwargs={"pk": self.lead.pk}),
            data={"close_status": "WYGRANA"},
        )
        self.assertEqual(response.status_code, 403)
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.status, Lead.Status.NOWY)

    def test_invalid_status_does_not_close_lead(self) -> None:
        """Niepoprawny status nie zmienia leada."""
        self.client.force_login(self.owner)
        self.client.post(
            reverse("leads:close", kwargs={"pk": self.lead.pk}),
            data={"close_status": "NIEPOPRAWNY"},
        )
        self.lead.refresh_from_db()
        self.assertEqual(self.lead.status, Lead.Status.NOWY)

    def test_close_redirects_to_detail(self) -> None:
        """Po zamknieciu nastepuje przekierowanie na szczegoly leada."""
        self.client.force_login(self.owner)
        response = self.client.post(
            reverse("leads:close", kwargs={"pk": self.lead.pk}),
            data={"close_status": "WYGRANA"},
        )
        self.assertRedirects(
            response, reverse("leads:detail", kwargs={"pk": self.lead.pk})
        )


# ---------------------------------------------------------------------------
# LeadExportView
# ---------------------------------------------------------------------------


class LeadExportViewTest(TestCase):
    """Testy eksportu leadow do XLSX."""

    def setUp(self) -> None:
        self.admin = _make_user("lexp_admin", role=UserProfile.Role.ADMIN)
        self.handlowiec = _make_user("lexp_hand")
        self.other = _make_user("lexp_other")
        self.company = _make_company("Exp Co", owner=self.handlowiec)
        stage = _get_stage()
        self.own_lead = Lead.objects.create(
            title="Moj Lead", company=self.company, owner=self.handlowiec, stage=stage
        )
        self.other_lead = Lead.objects.create(
            title="Cudzy Lead", company=self.company, owner=self.other, stage=stage
        )

    def test_export_redirect_anonymous(self) -> None:
        """Anonimowy uzytkownik jest przekierowywany do logowania."""
        response = self.client.get(reverse("leads:export_xlsx"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_export_returns_200_for_logged_user(self) -> None:
        """Zalogowany uzytkownik otrzymuje odpowiedz 200."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:export_xlsx"))
        self.assertEqual(response.status_code, 200)

    def test_export_content_type(self) -> None:
        """Odpowiedz ma Content-Type xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:export_xlsx"))
        self.assertIn("spreadsheetml.sheet", response["Content-Type"])

    def test_export_content_disposition_attachment(self) -> None:
        """Odpowiedz zwraca zalacznik z rozszerzeniem .xlsx."""
        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:export_xlsx"))
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_handlowiec_exports_only_own_leads(self) -> None:
        """HANDLOWIEC eksportuje tylko swoje leady."""
        import io as _io

        import openpyxl

        self.client.force_login(self.handlowiec)
        response = self.client.get(reverse("leads:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moj Lead", titles)
        self.assertNotIn("Cudzy Lead", titles)

    def test_admin_exports_all_leads(self) -> None:
        """ADMIN eksportuje wszystkie leady."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        titles = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Moj Lead", titles)
        self.assertIn("Cudzy Lead", titles)

    def test_export_has_header_row(self) -> None:
        """Plik XLSX zawiera wiersz naglowkowy z polem ID i Tytul."""
        import io as _io

        import openpyxl

        self.client.force_login(self.admin)
        response = self.client.get(reverse("leads:export_xlsx"))
        wb = openpyxl.load_workbook(_io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Tytuł")
